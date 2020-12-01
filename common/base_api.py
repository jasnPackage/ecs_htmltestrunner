# coding:utf-8
import json,xlrd,openpyxl
import requests,ast,datetime,re,time
from common.readexcel import ExcelUtil
from openpyxl.styles import Font,Alignment
from common.write_excel import copy_excel, Write_excel
from config import readConfig
from common import base_login,base_rely
from common.parameters_handle import parameters_handle
from common.mysql_pub import MysqlUtil
from common.logger import Log
from common.comm_function import *

mysql_db = MysqlUtil()
log = Log()
cookies = rely_func = ''
count_rows = 0


def send_requests(s, testdata):
    pass_case = 0  # 通过用例数标识
    fail_case = 0  # 失败用例数标识
    res = {}  # 接受返回数据
    # 接口地址
    interface_address = testdata["接口地址"]
    global cookies,rely_func,count_rows

    book = xlrd.open_workbook(testdata['用例路径'])
    '''
    使用openpyxl操作excel，openpyxl支持excel2007以上版本
    '''
    wb = openpyxl.load_workbook(testdata['用例路径'])
    ws = wb.worksheets[1]
    font_green = Font(color="37b400")
    font_red = Font(color="ff0000")

    '''
    获取excel文件中测试用例信息，将A-K列转换成字典存储
    '''
    sheet = book.sheet_by_index(1)  # 通过索引，获取相应的列表，这里表示获取excel的第二个列表
    ncols = sheet.ncols  # 获取所有列数

    '''
    获取cfg.ini配置文件中接口公共信息（ip和port）和run_config.ini配置文件中
    运行信息
    '''
    runmode = readConfig.runmode
    if runmode == 'test':
        ip = readConfig.test_ip  # 获取配置文件中接口ip
        port = readConfig.test_port  # 获取配置文件中接口port
    elif runmode == 'pro':
        ip = readConfig.pro_ip  # 获取配置文件中接口ip
        port = readConfig.pro_port  # 获取配置文件中接口port
    else:
        raise ValueError("run_config.ini文件运行环境(runmode)错误")

    url = "http://" + ip + ":" + port + interface_address
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64; rv:44.0) Gecko/20100101 Firefox/44.0",
        "X-Requested-With": "XMLHttpRequest",
        "Connection": "keep-alive"
    }

    # 读取测试用例的数据
    par = testdata["请求内容"]
    userlogin_token = testdata["是否需要登录token"]
    teardown_sql = testdata["teardown执行SQL"]
    rely_datas = testdata["依赖数据"]
    case_level = testdata["用例级别"]
    inf_mode = testdata["请求类型"]

    # 读取run_config.ini中设置用例运行级别
    runCase_level = readConfig.runCase_level

    if rely_datas:
        list_rely_data = ast.literal_eval(rely_datas)  # string转换成list
        for j in list_rely_data:
            if '/' in j:
                rely_data = j.split('/')
                relyexcel_name = rely_data[0] + ".xlsx"  # 依赖用例所在Excel名称
                relycase_id = rely_data[1]  # 依赖用例编号
                relycase_path = rely_data[2]  # 依赖数据的取值路径
                rely_value = base_rely.get_response_value(relyexcel_name, relycase_id, relycase_path)
                # 判断请求参数中是否需要userid或者其他变量
                if '$' in par:
                    par = parameters_handle(par, rely_value)
                else:
                    par = par
            elif '(' and ')' in j:
                rely_func = j

        if '*' in par:
            par = parameters_handle(par, eval(rely_func), partten=r"\*{(.*?)}")
        else:
            par = par

    # 判断测试用例是否需要登录的cookie
    if interface_address.find('/triage/') >= 0:
        if userlogin_token == '是':
            cookies = base_login.userlogin_cookies()
        else:
            cookies = ''
    elif interface_address.find('/triagedoctor/') >= 0:
        if userlogin_token == '是':
            cookies = base_login.triagedoctor_userlogin_cookies()
        else:
            cookies = ''
    elif interface_address.find('/triagerescue/') >= 0:
        if userlogin_token == '是':
            cookies = base_login.triagerescue_userlogin_cookies()
        else:
            cookies = ''
    elif interface_address.find('/triageobserve/') >= 0:
        if userlogin_token == '是':
            cookies = base_login.triageobserve_userlogin_cookies()
        else:
            cookies = ''



    # 统计每个测试用例的执行时间
    # 用例执行开始时间
    now_time = datetime.datetime.now()
    start_time = datetime.datetime.strptime(str(now_time), '%Y-%m-%d %H:%M:%S.%f')  # 测试用例执行开始时间

    '''
    判断请求方式是GET还是POST，并且判断测试用例预期结果与实际响应一致，所有接口请求前先调用登录接口
    '''

    if case_level == runCase_level or runCase_level == "ALL":
        if inf_mode == 'GET':
            r = s.get(url, params=par.encode("utf-8"), cookies=cookies)
            result = r.text
        elif inf_mode == 'POST':
            if testdata["请求头header"] == "form":
                headers["Content-Type"] = "application/x-www-form-urlencoded"
            elif testdata["请求头header"] == "json":
                headers["Content-Type"] = "application/json"
            else:
                raise ValueError("请求头header不能识别")
            r = s.post(url, data=par.encode("utf-8"), headers=headers, cookies=cookies)
            result = r.text
        res["实际响应"] = result

        # 执行teardown的SQL，如果有多个SQL，换行符分隔，循环执行
        if teardown_sql:
            exec_sqls = teardown_sql.split('\n')
            for exec_sql in exec_sqls:
                mysql_db.mysql_execute(exec_sql)

        # 先清空实际响应单元格的内容
        ws.cell(row=testdata['rowNum'], column=ncols-3, value='')
        # 设置实际响应写入的内容中对齐方式，居中对齐
        align = Alignment(horizontal='left', vertical='center', wrap_text=True)  # 纯色填充
        # 将实际响应写入测试用例文件中
        ws.cell(row=testdata['rowNum'], column=ncols-3, value=result).alignment = align

        log.info("\n用例路径：%s\n接口名称：%s\n用例编号：%s\n接口地址：%s\n用例名称：%s\n用例描述：%s\n请求内容：%s\n实际响应：%s" % (
        testdata["用例路径"], testdata["接口名称"], testdata["用例编号"], interface_address, testdata["用例名称"], testdata["用例描述"], par, result))

        checkPoints = testdata["检查点"]
        json_result = json.loads(result)
        if 'rows' in result and isinstance(json_result['rows'], list):
            count_rows = len(json_result['rows'])

        if checkPoints:
            list_checkPoints = ast.literal_eval(checkPoints)  # string转换成list

            # 设置检查点通过失败次数
            checkPoint_pass = 0
            checkPoint_fail = 0

            for checkPoint in list_checkPoints:
                # 将dict转换成json，并且去除转换成json后冒号后的空格
                json_checkPoint = json.dumps(checkPoint, separators=(',', ':'), ensure_ascii=False)
                # 通过正则取{}中间的值
                new_checkPoint = (re.findall('{(.*)}', json_checkPoint))[0]
                if 'rows' not in new_checkPoint:
                    if new_checkPoint in result:
                        checkPoint_pass += 1
                        log.info("检查点:%s--pass" % new_checkPoint)
                        log.info("----------------")
                    else:
                        checkPoint_fail += 1
                        log.error("检查点:%s--fail" % new_checkPoint)
                        log.info("----------------")
                else:
                    if int(new_checkPoint.split(':')[-1]) == count_rows:
                        checkPoint_pass += 1
                        log.info("检查点:%s--pass" % new_checkPoint)
                        log.info("----------------")
                    else:
                        checkPoint_fail += 1
                        log.error("检查点:%s--fail" % new_checkPoint)
                        log.info("----------------")
            log.info("\n\n")
            if checkPoint_pass == len(list_checkPoints):
                pass_case += 1
                ws.cell(row=testdata['rowNum'], column=ncols, value='Pass').font = font_green
                res["执行结果"] = "Pass"
            else:
                fail_case += 1
                ws.cell(row=testdata['rowNum'], column=ncols, value='Fail').font = font_red
                res["执行结果"] = "Fail"
        else:
            log.info("预置数据用例，不需要检查点")
            log.info("----------------")
            log.info("\n\n")
            res["执行结果"] = ''

        # 用例执行结束时间
        now_time = datetime.datetime.now()
        end_time = datetime.datetime.strptime(str(now_time), '%Y-%m-%d %H:%M:%S.%f')
        # 用例执行耗费时间
        elapse_time = str((end_time - start_time)) + "秒"
        res["执行耗时"] = elapse_time

        # 先清空执行耗时单元格的内容
        ws.cell(row=testdata['rowNum'], column=ncols-1, value='')
        # 将用例执行耗时写入测试用例中
        ws.cell(row=testdata['rowNum'], column=ncols-1, value=elapse_time).alignment = align

        # 最近一次用例执行时间
        execution_time = now_time.strftime('%Y-%m-%d %H:%M:%S')
        res["最近一次执行时间"] = execution_time
        ws.cell(row=testdata['rowNum'], column=ncols-2, value=execution_time).alignment = align

        # 保存测试用例文件
        wb.save(testdata['用例路径'])
    else:
        log.info("用例编号:%s--非指定运行级别的用例，不执行" % testdata["用例编号"])
    return res


