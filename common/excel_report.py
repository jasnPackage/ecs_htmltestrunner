#coding:utf-8
import base

from openpyxl import Workbook
from openpyxl.styles import Font,Alignment
import time,os,string
from config import readConfig


'''
获取excel测试报告所在的目录
'''
d = os.path.dirname(__file__) #返回当前文件所在目录（common文件夹路径）
parent_path = os.path.dirname(d) #返回common的父级目录
data_path = os.path.join(parent_path,'excel') #返回data所在目录
# print(data_path)
# data_path1 = os.listdir(data_path) #返回data目录下所有的文件

wb = Workbook()
ws = wb.active

def cell_style(row,column,value):
    font_2 = Font(name='Arial', size=16.5)  # 设置测试
    align = Alignment(horizontal='center', vertical='center')

    ws.cell(row=row,column=column,value=value).font = font_2
    ws.cell(row=row, column=column).alignment = align

def create_excel_report(test_time,pass_case,fail_case,success_rate):
    excel_name = time.strftime('%Y-%m-%d %H_%M_%S')
    # print(excel_name)

    ws.title = "测试结果"

    font_1 = Font(name='微软雅黑',size=21.5,bold=True) #设置“测试报告”单元格字体样式
    align = Alignment(horizontal='center',vertical='center')

    ws.cell(row=1,column=1).alignment = align
    ws.cell(row=1,column=1,value="测试报告").font = font_1

    pro_name = readConfig.pro_name
    pro_version = readConfig.pro_version
    tice_time = readConfig.tice_time
    tice_person = readConfig.tice_person
    tester = readConfig.tester
    auditor = readConfig.auditor

    cell_style(3,1,"测试详情")
    cell_style(4,1,"项目名称")
    cell_style(5,1,"接口版本")
    cell_style(6,1,"提测人")

    cell_style(4,2,pro_name)
    cell_style(5,2,pro_version)
    cell_style(6,2,tice_person)

    cell_style(4,3,"测试人")
    cell_style(5,3,"测试时间")
    cell_style(6,3,"审核人")

    cell_style(4,4,tester)
    cell_style(5,4,test_time)
    cell_style(6,4,auditor)

    cell_style(4,5,"通过")
    cell_style(5,5,"失败")
    cell_style(6,5,"成功率")

    cell_style(4,6,pass_case)
    cell_style(5,6,fail_case)
    cell_style(6,6,success_rate)



    '''
    合并1-3行A列和G列
    '''
    for i in range(1,4):
        ws.merge_cells('A%s:G%s'%(i,i))

    '''
    设置A到G列列宽
    '''
    for i in string.ascii_uppercase[0:7]:
        ws.column_dimensions[i].width = 25
    '''
    设置第1到8行高
    '''
    for i in range(1,9):
        ws.row_dimensions[i].height = 31.5

    wb.save(data_path + '\%s测试报告.xlsx' % excel_name)

# create_excel_report()

