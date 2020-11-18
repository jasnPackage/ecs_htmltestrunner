#coding:utf-8
import base

import xlrd,os,json,ast,re,threading,time
import requests
import openpyxl
from openpyxl.styles import Font
from xlrd import xldate_as_tuple
from config import readConfig
from common.logger import Log
# from case.user_login import Login
from openpyxl.styles import Alignment
from common.excel_report import create_excel_report
from common.parameters_handle import parameters_handle
from common.mysql_pub import MysqlUtil
from common import base_login,base_rely
from common.comm_function import *


'''
获取测试用例data所在的目录
'''
d = os.path.dirname(__file__) #返回当前文件所在目录（common文件夹路径）
parent_path = os.path.dirname(d) #返回common的父级目录
data_path = os.path.join(parent_path,'data') #返回data所在目录
# data_path1 = os.listdir(data_path) #返回data目录下所有的文件

case_path = get_excel_lists(data_path) #返回data目录下所有的excel文件
# case_path = [os.path.join(data_path, i) for i in os.listdir(data_path) if i[-5:] == '.xlsx']

s = requests.session()
# lon = Login(s)
log = Log()
mysql_db = MysqlUtil()


def ecs_case():
    pass_case = 0  #通过用例数标识
    fail_case = 0  #失败用例数标识

    for filename in case_path:
        book = xlrd.open_workbook(filename)

        '''
        使用xlwt操作excel，xlwt只支持excel2007以下版本
        '''

        # wb = copy(book)
        # ws = wb.get_sheet(1)

        '''
        使用openpyxl操作excel，openpyxl支持excel2007以上版本
        '''
        wb = openpyxl.load_workbook(os.path.join(data_path,filename))
        ws = wb.worksheets[1]
        font_green = Font(color="37b400")
        font_red = Font(color="ff0000")


        '''
        获取excel文件中接口信息
        '''
        table = book.sheet_by_index(0)        # 通过索引，获取相应的列表，这里表示获取excel的第一个列表
        inf_name = table.row_values(1)[0]     # 接口名称
        # inf_address = table.row_values(1)[1]  # 接口地址
        inf_mode = table.row_values(1)[2]     # 请求方式

        '''
        获取excel文件中测试用例信息，将A-K列转换成字典存储
        '''
        sheet = book.sheet_by_index(1)  # 通过索引，获取相应的列表，这里表示获取excel的第二个列表
        nrows = sheet.nrows  # 获取所有行数
        ncols = sheet.ncols  # 获取所有列数
        key = sheet.row_values(0)

        for i in range(1,nrows):
            value = sheet.row_values(i)
            case_dict = dict(zip(key,value))

            # 读取测试用例中的用例级别
            case_level = case_dict["用例级别"]
            # 读取run_config.ini中设置用例运行级别
            runCase_level = readConfig.runCase_level

            # 接口地址
            interface_address = case_dict["接口地址"]

            # 如果用例级别是PD，先预置初始化数据
            # if case_level == '预置数据':
            #     new_inf_address = case_dict["接口地址"]
            # else:
            #     new_inf_address = inf_address

            '''
            获取cfg.ini配置文件中接口公共信息（ip和port）和run_config.ini配置文件中
            运行信息
            '''
            runmode = readConfig.runmode
            if runmode == 'test':
                ip = readConfig.test_ip  # 获取配置文件中接口ip
                i_port = readConfig.test_port  # 获取配置文件中接口port
            elif runmode == 'pro':
                ip = readConfig.pro_ip  # 获取配置文件中接口ip
                i_port = readConfig.pro_port  # 获取配置文件中接口port
            else:
                raise ValueError("run_config.ini文件运行环境(runmode)错误")

            url = "http://" + ip + ":" + i_port + interface_address
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64; rv:44.0) Gecko/20100101 Firefox/44.0",
                "X-Requested-With": "XMLHttpRequest",
                "Connection": "keep-alive"
            }


            # 读取测试用例的数据
            par = case_dict["请求体body"]
            userlogin_token = case_dict["是否需要登录token"]
            teardown_sql = case_dict["teardown执行SQL"]
            rely_datas = case_dict["依赖数据"]

            if rely_datas:
                list_rely_data=ast.literal_eval(rely_datas)  # string转换成list
                for j in list_rely_data:
                    if '/' in j:
                        rely_data = j.split('/')
                        relyexcel_name = rely_data[0] + ".xlsx"  # 依赖用例所在Excel名称
                        relycase_id = rely_data[1]  # 依赖用例编号
                        relycase_path = rely_data[2]  # 依赖数据的取值路径
                        rely_value = base_rely.get_response_value(relyexcel_name, relycase_id, relycase_path)
                        # 判断请求参数中是否需要userid或者其他变量
                        if '$' in par:
                            par = parameters_handle(par, rely_value)
                        else:
                            par = par
                    elif '(' and ')' in j:
                        rely_func = j


                if '*' in par:
                    par = parameters_handle(par, eval(rely_func), partten=r"\*{(.*?)}")
                else:
                    par = par


            # 判断测试用例是否需要登录的cookie
            if interface_address.find('/triage/') >=0:
                if userlogin_token == '是':
                    cookies = base_login.userlogin_cookies()
                else:
                    cookies = ''
            elif interface_address.find('/triagedoctor/') >=0:
                if userlogin_token == '是':
                    cookies = base_login.triagedoctor_userlogin_cookies()
                else:
                    cookies = ''
            elif interface_address.find('/triagerescue/') >=0:
                if userlogin_token == '是':
                    cookies = base_login.triagerescue_userlogin_cookies()
                else:
                    cookies = ''
            elif interface_address.find('/triageobserve/') >=0:
                if userlogin_token == '是':
                    cookies = base_login.triageobserve_userlogin_cookies()
                else:
                    cookies = ''


            # 统计每个测试用例的执行时间
            # 用例执行开始时间
            now_time = datetime.datetime.now()
            start_time = datetime.datetime.strptime(str(now_time),'%Y-%m-%d %H:%M:%S.%f')  # 测试用例执行开始时间

            '''
            判断请求方式是GET还是POST，并且判断测试用例预期结果与实际响应一致，所有接口请求前先调用登录接口
            '''
            if case_level == runCase_level or runCase_level == "ALL":

                if inf_mode == 'GET':
                    r = s.get(url, params=par.encode("utf-8"),cookies=cookies)
                    result = r.text
                elif inf_mode == 'POST':
                    if case_dict["请求头header"] == "form":
                        headers["Content-Type"] = "application/x-www-form-urlencoded"
                    elif case_dict["请求头header"] == "json":
                        headers["Content-Type"] = "application/json"
                    else:
                        raise ValueError("请求头header不能识别")
                    r = s.post(url, data=par.encode("utf-8"), headers=headers,cookies=cookies)
                    result = r.text


                # 执行teardown的SQL，如果有多个SQL，换行符分隔，循环执行
                if  teardown_sql:
                    exec_sqls = teardown_sql.split('\n')
                    for exec_sql in exec_sqls:
                        mysql_db.mysql_execute(exec_sql)

                # 先清空实际响应单元格的内容
                ws.cell(row=i+1, column=ncols-3, value='')
                # 设置实际响应写入的内容中对齐方式，居中对齐
                align = Alignment(horizontal='left', vertical='center', wrap_text=True)  # 纯色填充
                # 将实际响应写入测试用例文件中
                ws.cell(row=i+1, column=ncols-3, value=result).alignment = align

                log.info("\n用例文件：%s\n用例编号：%s\n接口名称：%s\n接口地址：%s\n用例名称：%s\n用例描述：%s\n请求内容：%s\n实际响应：%s" % (
                filename,case_dict["用例编号"], inf_name, interface_address, case_dict["用例名称"], case_dict["用例描述"], par, result))

                checkPoints = case_dict["检查点"]
                json_result = json.loads(result)
                if 'rows' in result and isinstance(json_result['rows'],list):
                    count_rows = len(json_result['rows'])


                if checkPoints:
                    list_checkPoints = ast.literal_eval(checkPoints)  # string转换成list

                    # 设置检查点通过失败次数
                    checkPoint_pass = 0
                    checkPoint_fail = 0

                    for checkPoint in list_checkPoints:
                        # 将dict转换成json，并且去除转换成json后冒号后的空格
                        json_checkPoint = json.dumps(checkPoint, separators=(',', ':'), ensure_ascii=False)
                        # 通过正则取{}中间的值
                        new_checkPoint = (re.findall('{(.*)}', json_checkPoint))[0]
                        if 'rows' not in new_checkPoint:
                            if new_checkPoint in result:
                                checkPoint_pass += 1
                                log.info("检查点:%s--pass" % new_checkPoint)
                                log.info("----------------")
                            else:
                                checkPoint_fail += 1
                                log.error("检查点:%s--fail" % new_checkPoint)
                                log.info("----------------")
                        else:
                            if int(new_checkPoint.split(':')[-1]) == count_rows:
                                checkPoint_pass += 1
                                log.info("检查点:%s--pass" % new_checkPoint)
                                log.info("----------------")
                            else:
                                # ws.write(i,ncols-1,'fail'.encode('utf-8'))
                                # ws.cell(row=i + 1, column=ncols, value='Fail').font = font_red
                                checkPoint_fail += 1
                                log.error("检查点:%s--fail" % new_checkPoint)
                                log.info("----------------")

                    if checkPoint_pass == len(list_checkPoints):
                        pass_case += 1
                        ws.cell(row=i + 1, column=ncols, value='Pass').font = font_green
                    else:
                        fail_case += 1
                        ws.cell(row=i + 1, column=ncols, value='Fail').font = font_red

                # 用例执行结束时间
                now_time = datetime.datetime.now()
                end_time = datetime.datetime.strptime(str(now_time), '%Y-%m-%d %H:%M:%S.%f')
                # 用例执行耗费时间
                elapse_time = str((end_time - start_time)) + "秒"
                # 先清空执行耗时单元格的内容
                ws.cell(row=i + 1, column=ncols - 1, value='')

                # 将用例执行耗时写入测试用例中
                ws.cell(row=i + 1, column=ncols - 1, value=elapse_time).alignment = align


                # 最近一次用例执行时间
                execution_time = now_time.strftime('%Y-%m-%d %H:%M:%S')
                ws.cell(row=i + 1, column=ncols - 2, value=execution_time).alignment = align

                # 保存测试用例文件
                wb.save(os.path.join(data_path, filename))

            else:
                log.info("用例编号:%s--非指定运行级别的用例，不执行"%case_dict["用例编号"])

        total_case = pass_case + fail_case  # 执行用例总数
        log.info("执行测试用例数：%s  通过用例数：%s  失败用例数：%s" % (total_case, pass_case, fail_case))

    # 创建Excel测试报告
    create_excel_report(now_time, pass_case, fail_case, "%.2f%%" % (pass_case / (total_case) * 100))


ecs_case()