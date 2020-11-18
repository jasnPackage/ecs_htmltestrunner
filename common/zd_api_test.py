#coding:utf-8

import xlrd,os
import requests
import openpyxl
import unittest
from openpyxl.styles import Font
# from xlutils.copy import copy
from datetime import datetime
from xlrd import xldate_as_tuple
from config import readConfig
from common.logger import Log
from case.user_login import Login

'''
获取测试用例data所在的目录
'''
d = os.path.dirname(__file__) #返回当前文件所在目录（common文件夹路径）
parent_path = os.path.dirname(d) #返回common的父级目录
data_path = os.path.join(parent_path,'data') #返回data所在目录
data_path1 = os.listdir(data_path) #返回data目录下所有的文件

s = requests.session()
lon = Login(s)
log = Log()
a = []

class Api_test(unittest.TestCase):

    def __init__(self,str):
        self.str = str
        setattr(self.__class__, self.str, self.test_api_data)

class TestApi(Api_test):
    def test_api_data(self):
        p = 1  # 用例总数标识
        pa = 0  # 通过用例数标识
        fa = 0  # 失败用例数标识

        for filename in data_path1:
            book = xlrd.open_workbook(os.path.join(data_path, filename))

            '''
            使用xlwt操作excel，xlwt只支持excel2007以下版本
            '''

            # wb = copy(book)
            # ws = wb.get_sheet(1)

            '''
            使用openpyxl操作excel，openpyxl支持excel2007以上版本
            '''
            wb = openpyxl.load_workbook(os.path.join(data_path, filename))
            ws = wb.worksheets[1]
            font_green = Font(color="37b400")
            font_red = Font(color="ff0000")

            '''
            获取excel文件中接口信息
            '''
            table = book.sheet_by_index(0)  # 通过索引，获取相应的列表，这里表示获取excel的第一个列表
            inf_name = table.row_values(1)[0]  # 返回接口名称
            inf_address = table.row_values(1)[1]  # 返回接口地址
            inf_mode = table.row_values(1)[2]  # 返回请求方式

            '''
            获取excel文件中测试用例信息
            '''
            sheet = book.sheet_by_index(1)  # 通过索引，获取相应的列表，这里表示获取excel的第二个列表
            nrows = sheet.nrows  # 获取所有行数
            ncols = sheet.ncols  # 获取所有列数
            filed = sheet.row_values(0)
            # print(filed)

            for i in range(1, nrows):
                d1 = {}
                for j in range(0, len(filed) - 4):
                    ctype = sheet.cell(i, j).ctype  # 表格的数据类型
                    cell = sheet.cell_value(i, j)
                    d = {}
                    if ctype == 2 and cell % 1 == 0:  # 如果是整形
                        cell = int(cell)
                    elif ctype == 3:
                        # 转成datetime对象
                        date = datetime(*xldate_as_tuple(cell, 0))
                        cell = date.strftime('%Y/%m/%d')
                    elif ctype == 4:
                        cell = True if cell == 1 else False
                    # print(cell)
                    d.update({filed[j]: cell})
                    # print(d)
                    d1.update(d)
                # print(d1)

                '''
                获取excel文件中测试用例预期结果和描述
                '''

                for k in range(len(filed) - 4, len(filed) - 1):
                    ctype = sheet.cell(i, k).ctype  # 表格的数据类型
                    cell = sheet.cell_value(i, k)
                    if ctype == 2 and cell % 1 == 0:  # 如果是整形
                        cell = int(cell)
                    elif ctype == 3:
                        # 转成datetime对象
                        date = datetime(*xldate_as_tuple(cell, 0))
                        cell = date.strftime('%Y/%m/%d')
                    elif ctype == 4:
                        cell = True if cell == 1 else False
                    a.append(cell)
                # print(a)
                # print(type(a[0]))

                '''
                获取cfg.ini配置文件中接口公共信息（ip和port）和run_config.ini配置文件中
                运行信息
                '''
                runmode = readConfig.runmode
                if runmode == 'test':
                    ip = readConfig.test_ip  # 获取配置文件中接口ip
                    i_port = readConfig.test_port  # 获取配置文件中接口port
                elif runmode == 'pro':
                    ip = readConfig.pro_ip  # 获取配置文件中接口ip
                    i_port = readConfig.pro_port  # 获取配置文件中接口port
                else:
                    raise ValueError("run_config.ini文件运行环境(runmode)错误")

                url = "http://" + ip + ":" + i_port + inf_address
                headers = {
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64; rv:44.0) Gecko/20100101 Firefox/44.0",
                    "X-Requested-With": "XMLHttpRequest",
                    "Connection": "keep-alive"
                }
                par = d1

                '''
                判断请求方式是GET还是POST，并且判断测试用例预期结果与实际响应一致，所有接口请求前先调用登录接口
                '''
                code = "xuxingan"
                passwd = "admin"
                lon.login(code, passwd)

                if inf_mode == 'GET':
                    r = s.get(url, params=par)
                    result = r.json()
                    log.info("编号：%s，接口名称：%s，测试点：%s，响应：%s" % (p, inf_name, a[2], str(result).replace('None', 'null')))
                    # print(inf_name, str(result).replace('None','null'), a[1])
                    if str(result).replace('None', 'null') == a[0]:
                        # ws.write(i,ncols-1,'pass'.encode('utf-8'))
                        ws.cell(row=i + 1, column=ncols, value='pass').font = font_green
                        pa = pa + 1
                        log.info("pass")
                        log.info("--------")
                    else:
                        # ws.write(i,ncols-1,'fail'.encode('utf-8'))
                        ws.cell(row=i + 1, column=ncols, value='fail').font = font_red
                        fa = fa + 1
                        log.info("fail")
                        log.info("--------")
                elif inf_mode == 'POST':
                    r = s.post(url, data=par, headers=headers)
                    result = r.json()
                    log.info("编号：%s，接口名称：%s，测试点：%s，响应：%s" % (p, inf_name, a[2], str(result).replace('None', 'null')))
                    # print(inf_name, str(result).replace('None', 'null'), a[1])
                    if str(result).replace('None', 'null') == a[0]:
                        # ws.write(i,ncols-1,'pass'.encode('utf-8'))
                        ws.cell(row=i + 1, column=ncols, value='pass').font = font_green
                        pa = pa + 1
                        log.info("pass")
                        log.info("--------")
                    else:
                        # ws.write(i,ncols-1,'fail'.encode('utf-8'))
                        ws.cell(row=i + 1, column=ncols, value='fail').font = font_red
                        fa = fa + 1
                        log.info("fail")
                        log.info("--------")
                wb.save(os.path.join(data_path, filename))
                p = p + 1

        log.info("测试用例数：%s  成功用例数：%s  失败用例数：%s" % (p - 1, pa, fa))


if __name__ == '__main__':
    a = Api_test(a[1])
    a.a[1]()
    unittest.main()